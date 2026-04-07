"""
clean_schedule.py — 数据清洗与归档脚本
════════════════════════════════════════

【核心功能】
  对晚自修排班 Excel 文件进行清洗、统计和归档，为求解器提供准确的历史数据基础。
  每次确认新月份排班无误、在 Excel 中完成改名后运行本脚本。

【核心逻辑】
  Step 1 · 修正"次数"sheet
    - 楼层字段标准化：2楼/3楼 → 2-3楼，4楼/5楼 → 4-5楼
    - 清除"要求"列中残留的"班主任"字样（防止重复追加）

  Step 2 · 统计历史值班次数
    - 只扫描 sheet 名格式严格为 "YYYY年M月" 的 sheet（如 2025年9月、2026年3月）
    - 兼容两种 sheet 内容格式：
        旧格式：含"星期"列，值为"一""二"…"日"
        新格式：含"日期"列，值为"5月6日(周二)"
    - 对"全体班主任"特殊标记，自动展开为全员计数

  Step 2.5 · 同步状态文件
    - 读取 scheduling_state.json 中的 generated_months
    - 将 Excel 中已有数据但未记录的月份自动补充，并更新 last_generated

  Step 3 · 重建"清洗后数据"sheet
    - 优先保留已有"清洗后数据"的 A–G 列（人工编辑内容不被覆盖）
    - 重新计算 H–M 列：历史总次数 / 周五次数 / 周日次数 / 三项月均值

【使用方法】
  1. 在 Excel 中将"2026年5月排班（暂定）"重命名为"2026年5月"
  2. 打开终端，运行：
       python clean_schedule.py
  3. 检查输出日志，确认月份统计和状态文件更新正确

  ⚠ 老师的增删、楼层、要求等基本资料请直接在"次数"sheet 中手动修改，
     本脚本不再自动补录老师。
"""

import json
import pathlib
import re
import sys
import openpyxl
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

_STATE_PATH = pathlib.Path(__file__).parent / "scheduling_state.json"

# 匹配 "YYYY年M月" 或 "YYYY年MM月" 格式的 sheet 名
_MONTHLY_SHEET_RE = re.compile(r'^\d{4}年\d{1,2}月$')

# ── 配置 ────────────────────────────────────────────────────────────────────
EXCEL_PATH = r"c:/Users/31249/Desktop/晚自修排版.xlsx"

BANZHUREN = {
    "胡晓娴", "丁亚男", "蒋寅", "许仲", "叶云", "吴徐帆",
    "陆子辰", "陆梦霞", "王仕全", "金丹萍", "毛建伟", "蓝申剑",
    "熊颖", "钟靖", "杨朋昊",
    "肖中海",
}

def is_monthly_sheet(name: str) -> bool:
    """返回True表示该sheet是合法的历史月份数据（格式：YYYY年M月）。"""
    return bool(_MONTHLY_SHEET_RE.match(name))


def sheet_name_to_month_key(name: str) -> str | None:
    """将 "2025年9月" 转换为 "2025-09"，不匹配则返回 None。"""
    m = _MONTHLY_SHEET_RE.match(name)
    if not m:
        return None
    year, month = name.split("年")
    month = month.rstrip("月")
    return f"{year}-{int(month):02d}"

FLOOR_MAP = {
    "2楼": "2-3楼",
    "3楼": "2-3楼",
    "4楼": "4-5楼",
    "5楼": "4-5楼",
}

VALID_DAYS = {"一", "二", "三", "四", "五", "六", "日"}


def extract_day(raw) -> str | None:
    """统一处理'星期三'和'三'两种格式，返回单字或 None。"""
    if raw is None:
        return None
    s = str(raw).strip()
    if s.startswith("星期") and len(s) == 3:
        return s[-1]
    return s if s in VALID_DAYS else None

# ── 辅助函数 ─────────────────────────────────────────────────────────────────

def normalize_floor(floor: str | None) -> str | None:
    if floor is None:
        return None
    return FLOOR_MAP.get(str(floor).strip(), str(floor).strip())


def find_header_row_idx(ws) -> int | None:
    """返回包含"星期"或"日期"的行号（1-based）。兼容旧格式和solver新格式。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in ("星期", "日期"):
                return cell.row
    return None


def col_index_by_name(ws, header_row_idx: int, name: str) -> int | None:
    """在 header_row_idx 行中找到值为 name 的列号（1-based）。"""
    for cell in ws[header_row_idx]:
        if cell.value == name:
            return cell.column
    return None


# ── Step 1：修改"次数"sheet ──────────────────────────────────────────────────

def patch_cishu_sheet(ws_cishu: openpyxl.worksheet.worksheet.Worksheet):
    for row in ws_cishu.iter_rows(min_row=2):
        name_cell   = row[1]   # Col B：姓名
        yaoqiu_cell = row[4]   # Col E：要求
        floor_cell  = row[5]   # Col F：楼层

        if not name_cell.value:
            continue

        # 楼层标准化
        floor_cell.value = normalize_floor(floor_cell.value)

        # 清除要求列中可能已写入的"班主任"标签（防止重复追加）
        if yaoqiu_cell.value and "班主任" in str(yaoqiu_cell.value):
            cleaned = "，".join(
                p for p in str(yaoqiu_cell.value).split("，") if p.strip() != "班主任"
            )
            yaoqiu_cell.value = cleaned if cleaned else None

    print("次数 sheet 修改完成（楼层标准化、班主任标签清理）。")


# ── Step 2：统计月份排班 ──────────────────────────────────────────────────────

def extract_day_from_date_label(val) -> str | None:
    """从 solver 新格式'5月6日(周二)'中提取星期字符'二'。"""
    if val is None:
        return None
    s = str(val).strip()
    if "(" in s and ")" in s:
        inner = s[s.index("(") + 1: s.index(")")]  # "周二"
        if inner.startswith("周") and len(inner) == 2 and inner[-1] in VALID_DAYS:
            return inner[-1]
    return None


def collect_duty_counts(wb) -> tuple[dict, dict, dict, int, list[str]]:
    total_count   = defaultdict(int)
    friday_count  = defaultdict(int)
    sunday_count  = defaultdict(int)
    active_months = 0
    active_list   = []

    monthly_sheets = [s for s in wb.sheetnames if is_monthly_sheet(s)]

    for sheet_name in monthly_sheets:
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        hdr_row = find_header_row_idx(ws)
        if hdr_row is None:
            print(f"  跳过（未找到'星期'或'日期'列）：{sheet_name}")
            continue

        # 支持旧格式（"星期"列）和新格式（"日期"列，值如"5月6日(周二)"）
        col_day = col_index_by_name(ws, hdr_row, "星期")
        use_date_format = False
        if col_day is None:
            col_day = col_index_by_name(ws, hdr_row, "日期")
            use_date_format = True

        col_1    = col_index_by_name(ws, hdr_row, "1楼")
        col_23   = col_index_by_name(ws, hdr_row, "2-3楼")
        col_45   = col_index_by_name(ws, hdr_row, "4-5楼")

        floor_cols = [c for c in [col_1, col_23, col_45] if c is not None]

        has_data = False
        for row in ws.iter_rows(min_row=hdr_row + 1):
            raw_day = row[col_day - 1].value if col_day else None
            if use_date_format:
                day = extract_day_from_date_label(raw_day)
            else:
                day = extract_day(raw_day)
            if day is None:
                continue

            is_fri = (day == "五")
            is_sun = (day == "日")
            has_data = True

            names: list[str] = []
            for col_idx in floor_cols:
                val = row[col_idx - 1].value
                if val:
                    names.append(str(val).strip())

            for name in names:
                if name == "全体班主任":
                    for bzr in BANZHUREN:
                        total_count[bzr] += 1
                        if is_fri:
                            friday_count[bzr] += 1
                        if is_sun:
                            sunday_count[bzr] += 1
                else:
                    total_count[name] += 1
                    if is_fri:
                        friday_count[name] += 1
                    if is_sun:
                        sunday_count[name] += 1

        if has_data:
            active_months += 1
            active_list.append(sheet_name)

    print(f"有数据的月份（{active_months}个）：{active_list}")
    return total_count, friday_count, sunday_count, active_months, active_list


# ── Step 2.5：同步 scheduling_state.json ────────────────────────────────────

def sync_state_file(active_list: list[str]) -> None:
    """
    将 Excel 中已有数据的月份 sheet（格式 YYYY年M月）同步到 scheduling_state.json。
    若某月已在 Excel 中归档但未记录在 generated_months，则自动补充。
    """
    if not _STATE_PATH.exists():
        print("  scheduling_state.json 不存在，跳过状态同步。")
        return

    state = json.loads(_STATE_PATH.read_text(encoding="utf-8"))
    generated: list[str] = state.get("generated_months", [])

    added: list[str] = []
    for sheet_name in active_list:
        key = sheet_name_to_month_key(sheet_name)
        if key and key not in generated:
            generated.append(key)
            added.append(f"{sheet_name} → {key}")

    if added:
        generated.sort()
        state["generated_months"] = generated
        state["last_generated"] = generated[-1]
        _STATE_PATH.write_text(
            json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"  scheduling_state.json 补充了 {len(added)} 个月份：")
        for a in added:
            print(f"    {a}")
        print(f"  last_generated 更新为：{state['last_generated']}")
    else:
        print("  scheduling_state.json 无需更新（所有月份已记录）。")


# ── Step 3：生成"清洗后数据"sheet ─────────────────────────────────────────────

def build_clean_sheet(
    wb,
    ws_cishu,
    total_count: dict,
    friday_count: dict,
    sunday_count: dict,
    active_months: int,
):
    # 先把现有"清洗后数据"的 A-F 列读出来缓存（保留人工编辑内容）
    # 键：姓名；值：(xueke, nianji, yaoqiu, floor) — 序号重新按顺序编
    preserved_af: dict[str, tuple] = {}
    if "清洗后数据" in wb.sheetnames:
        for row in wb["清洗后数据"].iter_rows(min_row=2, values_only=True):
            if row[1]:  # 姓名不为空
                name = str(row[1]).strip()
                preserved_af[name] = (row[2], row[3], row[4], row[5], row[6])  # 学科,年级,要求,楼层,是否班主任
        del wb["清洗后数据"]
        print(f"  已读取 {len(preserved_af)} 条 A-G 列缓存（人工编辑内容将被保留）。")

    # 插入在"次数"之后
    cishu_idx = wb.sheetnames.index(wb.sheetnames[0])
    ws_clean = wb.create_sheet("清洗后数据", cishu_idx + 1)

    denom_label = f"/{active_months}月" if active_months else ""
    headers = [
        "序号", "姓名", "教学学科", "教学年级", "要求", "楼层", "是否班主任",
        "历史总次数", "历史周五次数", "历史周日次数",
        f"月均次数{denom_label}", f"月均周五{denom_label}", f"月均周日{denom_label}",
    ]
    ws_clean.append(headers)

    seq = 1
    for row in ws_cishu.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name:
            continue
        name = str(name).strip()

        if name in preserved_af:
            # 优先使用"清洗后数据"中保留的 A-G 值（人工编辑内容）
            xueke, nianji, yaoqiu, floor, is_bzr = preserved_af[name]
        else:
            # 新增老师：从"次数"sheet 读取，是否班主任由 BANZHUREN 集合决定
            xueke  = row[2]
            nianji = row[3]
            yaoqiu = row[4]
            floor  = row[5]
            is_bzr = "是" if name in BANZHUREN else "否"

        total   = total_count.get(name, 0)
        friday  = friday_count.get(name, 0)
        sunday  = sunday_count.get(name, 0)

        if active_months > 0:
            avg_total = round(total  / active_months, 2)
            avg_fri   = round(friday / active_months, 2)
            avg_sun   = round(sunday / active_months, 2)
        else:
            avg_total = avg_fri = avg_sun = 0

        ws_clean.append([
            seq, name, xueke, nianji, yaoqiu, floor, is_bzr,
            total, friday, sunday, avg_total, avg_fri, avg_sun,
        ])
        seq += 1

    print(f"清洗后数据 sheet 写入完成，共 {seq - 1} 行。")


# ── 主入口 ────────────────────────────────────────────────────────────────────

def main():
    print(f"加载文件：{EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    cishu_name = wb.sheetnames[0]
    print(f"次数 sheet 名：{cishu_name}")
    ws_cishu = wb[cishu_name]

    print("\n── Step 1：修改次数 sheet ──")
    patch_cishu_sheet(ws_cishu)

    print("\n── Step 2：统计月份排班 ──")
    total_count, friday_count, sunday_count, active_months, active_list = collect_duty_counts(wb)

    print("\n── Step 2.5：同步 scheduling_state.json ──")
    sync_state_file(active_list)

    print("\n── Step 3：生成清洗后数据 ──")
    build_clean_sheet(wb, ws_cishu, total_count, friday_count, sunday_count, active_months)

    print(f"\n保存文件：{EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    print("完成！")


if __name__ == "__main__":
    main()
