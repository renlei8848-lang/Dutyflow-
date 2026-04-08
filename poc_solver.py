"""
poc_solver.py — CP-SAT 排班求解器
════════════════════════════════════════

【核心功能】
  接收教师数据列表，基于真实日历构建约束规划模型，求解最优值班方案，
  并将结果导出到 Excel 文件的指定 sheet。

【核心逻辑】
  日历构建（build_schedule_days）
    · 枚举目标月所有日期，保留周一至周五和周日为值班日
    · 按 special_dates.json 配置：剔除节假日区间、加入调休周六、
      在全体班主任活动日抑制班主任的变量创建
    · 按 ISO 周分组，分配 week_idx（用于周次上限约束）

  变量矩阵（_create_variables）
    · 稀疏 BoolVar 矩阵，键为 (teacher_id, week_idx, day_of_week, slot_id)
    · 仅为楼层匹配、标签合法的教师创建变量

  硬约束（HC-1 ~ HC-6）
    · HC-1 覆盖率：每个值班日每个楼层恰好1人
    · HC-2 不重排：同一教师同一天最多1个楼层
    · HC-3 周次上限：每位教师每周最多1次
    · HC-4 月次上限：每位教师每月最多2次（标注"一个月只能1次"者限1次）
    · HC-5 周末互斥：周五和周日合计最多1次
    · HC-6 极差约束：同组教师（班主任/非班主任）本月新排次数极差 ≤ 1

  软约束目标函数（SC-1 ~ SC-5，最大化）
    · 学科/星期偏好奖励（班主任排周末、英语排周一/四、语文排周二、政治排周三）
    · 班主任排非周末惩罚（SC-4）
    · 非班主任多排奖励、间隔拉开奖励（SC-1/SC-2）
    · 偏离月均目标惩罚（SC-3）
    · 非班主任周末双排惩罚（SC-5）

  输出
    · Sheet 名格式：{year}年{month}月排班（暂定），如"2026年5月排班（暂定）"
    · 日期列格式："5月6日(周二)"
    · 班主任单元格橙黄色高亮

【使用方法】
  通常由 main.py 调用，无需单独运行。
  若需单独测试，可直接执行：
      python poc_solver.py
  或指定 Excel 路径和目标月：
      python poc_solver.py "C:/path/to/晚自修排版.xlsx" "2026-05"
  结果写入 Excel，调试报告写入 debug_solver_run.txt。

  模块边界（不可跨越）：
    · 本模块不读取 Excel，仅接收 List[TeacherRecord]
    · 所有数值常量来自 PARAMS_REGISTRY.yaml
    · 运行时不调用任何 LLM/AI 接口
"""

from __future__ import annotations

import calendar
import datetime
import json
import pathlib
from collections import namedtuple
from datetime import date, timedelta

import pandas as pd
import yaml
from openpyxl.styles import Alignment, PatternFill
from ortools.sat.python import cp_model

from poc_loader import TeacherRecord, load_teacher_records

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------
_REGISTRY_PATH = pathlib.Path(__file__).parent / ".dutyflow_meta" / "PARAMS_REGISTRY.yaml"
_STATE_PATH = pathlib.Path(__file__).parent / "scheduling_state.json"
_SPECIAL_DATES_PATH = pathlib.Path(__file__).parent / "special_dates.json"

with _REGISTRY_PATH.open(encoding="utf-8") as _f:
    _CFG = yaml.safe_load(_f)

_DEFAULT_EXCEL_PATH: str = _CFG["clean_schedule"]["excel_path"]

# slot_id → floor_zone string (must match TeacherRecord.floor_zone)
SLOT_ZONE: dict[str, str] = {
    "floor_1":   "1楼",
    "floor_2_3": "2-3楼",
    "floor_4_5": "4-5楼",
}
SLOT_IDS: list[str] = list(SLOT_ZONE.keys())

# Day-of-week index → display name (Python weekday: 0=Mon … 4=Fri, 5=Sat, 6=Sun)
DAY_NAMES: dict[int, str] = {
    0: "周一", 1: "周二", 2: "周三", 3: "周四", 4: "周五", 5: "周六", 6: "周日",
}

# Base duty weekdays per Python weekday() convention (Mon–Fri + Sun, excluding Sat)
_BASE_DUTY_WEEKDAYS: frozenset[int] = frozenset({0, 1, 2, 3, 4, 6})

# ---------------------------------------------------------------------------
# Calendar helper
# ---------------------------------------------------------------------------
ScheduleDay = namedtuple("ScheduleDay", ["date", "week_idx", "day_of_week"])
"""
A single duty day in the scheduling cycle.
  date        — actual calendar date (datetime.date)
  week_idx    — 0-based sequential week index within this month (drives HC-3)
  day_of_week — Python weekday int (0=Mon … 4=Fri, 5=Sat, 6=Sun)
"""


def build_schedule_days(year: int, month: int, special_cfg: dict) -> list[ScheduleDay]:
    """
    Compute all duty days for a given (year, month) given special-date configuration.

    special_cfg keys (all optional):
      holiday_ranges       — list of [start_str, end_str] date ranges (inclusive); removed from duty days
      extra_workdays       — list of date strings for makeup workdays (Saturdays); added to duty days
      all_bz_required_days — dates where BZ teachers must attend; suppressed in _create_variables

    Returns a list of ScheduleDay sorted by date.
    """
    # Parse holiday date ranges
    holiday_dates: set[date] = set()
    for pair in special_cfg.get("holiday_ranges", []):
        cur = date.fromisoformat(pair[0])
        end = date.fromisoformat(pair[1])
        while cur <= end:
            holiday_dates.add(cur)
            cur += timedelta(days=1)

    # Parse extra workdays (makeup Saturdays)
    extra_workday_dates: set[date] = {
        date.fromisoformat(s) for s in special_cfg.get("extra_workdays", [])
    }

    # Enumerate all duty dates for this month
    _, num_days = calendar.monthrange(year, month)
    duty_dates: list[date] = []
    seen: set[date] = set()

    for day_num in range(1, num_days + 1):
        d = date(year, month, day_num)
        if d in holiday_dates:
            continue
        if d.weekday() in _BASE_DUTY_WEEKDAYS or d in extra_workday_dates:
            duty_dates.append(d)
            seen.add(d)

    # Extra workdays might be Saturdays (not in _BASE_DUTY_WEEKDAYS); add if not already included
    for d in sorted(extra_workday_dates):
        if d.year == year and d.month == month and d not in seen and d not in holiday_dates:
            duty_dates.append(d)
    duty_dates.sort()

    # Group by ISO (year, week) → assign sequential week_idx
    iso_weeks: list[tuple[int, int]] = []
    iso_seen: set[tuple[int, int]] = set()
    for d in duty_dates:
        yw: tuple[int, int] = d.isocalendar()[:2]
        if yw not in iso_seen:
            iso_weeks.append(yw)
            iso_seen.add(yw)
    week_idx_map: dict[tuple[int, int], int] = {yw: i for i, yw in enumerate(iso_weeks)}

    return [
        ScheduleDay(
            date=d,
            week_idx=week_idx_map[d.isocalendar()[:2]],
            day_of_week=d.weekday(),
        )
        for d in duty_dates
    ]


def _next_month(last: str) -> str:
    """Increment a 'YYYY-MM' string by one month."""
    year, month = map(int, last.split("-"))
    month += 1
    if month > 12:
        month = 1
        year += 1
    return f"{year:04d}-{month:02d}"


# ---------------------------------------------------------------------------
# Solver class
# ---------------------------------------------------------------------------
class DutySolver:
    def __init__(
        self,
        records: list[TeacherRecord],
        config: dict,
        target_month: str,
        special_dates: dict | None = None,
    ) -> None:
        """
        target_month  : "YYYY-MM" string for the scheduling cycle (e.g. "2026-05").
        special_dates : dict from special_dates.json for this month, or {} / None.
        """
        self._records = records
        self._config = config
        self._by_id: dict[int, TeacherRecord] = {t.teacher_id: t for t in records}
        self._target_month = target_month

        # Config extraction
        self._headcount: dict[str, int] = {
            sid: config["slots"][sid]["required_headcount"] for sid in SLOT_IDS
        }
        self._weights: dict[str, int] = config["weights"]
        self._time_limit: float = float(config["solver"]["time_limit_seconds"])
        self._num_workers: int = int(config["solver"]["num_search_workers"])
        self._seed: int = int(config["solver"]["random_seed"])
        self._log_progress: bool = bool(config["solver"]["log_search_progress"])

        mt = config["monthly_targets"]
        self._bz_target: int = int(mt["banzhuren_target"])
        self._non_bz_target: int = int(mt["non_banzhuren_target"])
        self._max_projected_range: int = int(mt["max_projected_range"])
        self._min_week_gap: int = int(config["soft_constraints"]["min_week_gap"])

        # Calendar: build real duty days for this month
        special_cfg: dict = special_dates or {}
        year, month = map(int, target_month.split("-"))
        self._schedule_days: list[ScheduleDay] = build_schedule_days(year, month, special_cfg)
        self._num_weeks: int = len({sd.week_idx for sd in self._schedule_days})

        # Fast lookup: (week_idx, day_of_week) → actual date
        self._date_by_wd: dict[tuple[int, int], date] = {
            (sd.week_idx, sd.day_of_week): sd.date for sd in self._schedule_days
        }

        # Dates where all BZ teachers must attend (duty variable suppressed for BZ on these dates)
        self._all_bz_days: set[date] = {
            date.fromisoformat(s) for s in special_cfg.get("all_bz_required_days", [])
        }

        # Number of mandatory BZ days falling in the target month.
        # When > 0, activates the two-system branch in _create_variables / HC-1 / HC-4 / SC-3 / SC-6.
        # When == 0, all original code paths are taken unchanged.
        self._mandatory_bz_count: int = sum(
            1 for d in self._all_bz_days
            if d.year == year and d.month == month
        )

        # Output sheet name, e.g. "2026年5月排班（暂定）"
        sheet_fmt: str = config["output"]["output_sheet_name_format"]
        self._output_sheet_name: str = (
            sheet_fmt.replace("{year}", str(year)).replace("{month}", str(month))
        )

        # CP-SAT objects (populated by _build)
        self.model = cp_model.CpModel()
        self.x: dict[tuple, cp_model.IntVar] = {}

        # Per-teacher IntVars (populated by _precompute_teacher_totals)
        self._new_total: dict[int, cp_model.IntVar] = {}
        self._new_friday: dict[int, cp_model.IntVar] = {}
        self._new_sunday: dict[int, cp_model.IntVar] = {}

        self.cp_solver: cp_model.CpSolver | None = None
        self.solver_status: int | None = None

        self._build()

    # -----------------------------------------------------------------------
    # Internal helpers
    # -----------------------------------------------------------------------
    def _date_label(self, w: int, d: int) -> str:
        """Return a human-readable label like '5月6日(周二)' for a (week_idx, day_of_week) pair."""
        actual = self._date_by_wd.get((w, d))
        if actual is None:
            return f"第{w + 1}周{DAY_NAMES.get(d, str(d))}"
        return f"{actual.month}月{actual.day}日({DAY_NAMES[d]})"

    # -----------------------------------------------------------------------
    # Internal: model construction
    # -----------------------------------------------------------------------
    def _build(self) -> None:
        self._create_variables()
        self._precompute_teacher_totals()
        self._add_hard_constraints()
        self._add_range_constraints()
        self._set_objective()

    def _create_variables(self) -> None:
        """Sparse variable matrix: only create BoolVars for valid assignments."""
        for t in self._records:
            if "不排了" in t.tags:
                continue
            for sd in self._schedule_days:
                w = sd.week_idx
                d = sd.day_of_week
                # Mandatory-day variable suppression (two-system branch):
                if self._mandatory_bz_count > 0:
                    # New path: suppress ALL teachers on mandatory days (全体班主任 covers the duty)
                    if sd.date in self._all_bz_days:
                        continue
                    # Hard-exclude BZ from Fri/Sun when mandatory days exist.
                    # Soft penalty proved insufficient: SC-5 savings for non-BZ (+150)
                    # outweighed the weekend penalty (-110 net), causing solver to
                    # place BZ on weekends. Hard exclusion within this branch is cleaner.
                    if t.is_banzhuren and d in (4, 6):
                        continue
                else:
                    # Original path: suppress only BZ teachers (兼容旧行为)
                    if t.is_banzhuren and sd.date in self._all_bz_days:
                        continue
                if d == 4 and "不要周五" in t.tags:
                    continue
                if d == 6 and "不要周日" in t.tags:
                    continue
                for slot_id in SLOT_IDS:
                    if t.floor_zone != SLOT_ZONE[slot_id]:
                        continue
                    key = (t.teacher_id, w, d, slot_id)
                    self.x[key] = self.model.NewBoolVar(
                        f"x_{t.teacher_id}_{w}_{d}_{slot_id}"
                    )

    def _precompute_teacher_totals(self) -> None:
        """
        For each active teacher (has at least one x-variable), create IntVars summing
        their new assignments. Reused by _add_range_constraints and _set_objective.
        """
        for t in self._records:
            tid = t.teacher_id
            all_vars = [v for (i, *_), v in self.x.items() if i == tid]
            if not all_vars:
                continue

            raw_cap = 1 if "一个月只能1次" in t.tags else 2
            # Two-system: when mandatory BZ days exist, BZ teachers' solver quota is reduced
            # by mandatory_bz_count (those days already count as duties).
            if self._mandatory_bz_count > 0 and t.is_banzhuren:
                solver_cap = max(0, raw_cap - self._mandatory_bz_count)
            else:
                solver_cap = raw_cap  # original path

            new_total = self.model.NewIntVar(0, solver_cap, f"new_total_{tid}")
            self.model.Add(new_total == sum(all_vars))
            self._new_total[tid] = new_total

            # Friday vars (d == 4)
            fri_vars = [v for (i, w, d, s), v in self.x.items() if i == tid and d == 4]
            new_friday = self.model.NewIntVar(0, solver_cap, f"new_fri_{tid}")
            self.model.Add(new_friday == (sum(fri_vars) if fri_vars else 0))
            self._new_friday[tid] = new_friday

            # Sunday vars (d == 6)
            sun_vars = [v for (i, w, d, s), v in self.x.items() if i == tid and d == 6]
            new_sunday = self.model.NewIntVar(0, solver_cap, f"new_sun_{tid}")
            self.model.Add(new_sunday == (sum(sun_vars) if sun_vars else 0))
            self._new_sunday[tid] = new_sunday

    def _add_hard_constraints(self) -> None:
        # HC-1: Coverage — every duty day × slot must have exactly required_headcount teachers
        for sd in self._schedule_days:
            # Mandatory BZ days have no solver variables (全体班主任 covers collectively);
            # skip coverage constraint for those days to avoid empty-slot RuntimeError.
            if self._mandatory_bz_count > 0 and sd.date in self._all_bz_days:
                continue
            w, d = sd.week_idx, sd.day_of_week
            for slot_id in SLOT_IDS:
                slot_vars = [
                    self.x[(tid, w, d, slot_id)]
                    for tid in self._by_id
                    if (tid, w, d, slot_id) in self.x
                ]
                if not slot_vars:
                    raise RuntimeError(
                        f"No eligible teachers for slot '{slot_id}' "
                        f"on {self._date_label(w, d)}. "
                        "Check floor assignments and tags in 清洗后数据 sheet."
                    )
                self.model.Add(sum(slot_vars) == self._headcount[slot_id])

        # HC-2: No-clone — each teacher at most 1 slot per day
        for t in self._records:
            for sd in self._schedule_days:
                w, d = sd.week_idx, sd.day_of_week
                day_vars = [
                    self.x[(t.teacher_id, w, d, s)]
                    for s in SLOT_IDS
                    if (t.teacher_id, w, d, s) in self.x
                ]
                if day_vars:
                    self.model.Add(sum(day_vars) <= 1)

        # HC-3: Weekly cap — each teacher at most 1 duty per week
        for t in self._records:
            for w in range(self._num_weeks):
                week_vars = [
                    self.x[(t.teacher_id, w, sd.day_of_week, s)]
                    for sd in self._schedule_days if sd.week_idx == w
                    for s in SLOT_IDS
                    if (t.teacher_id, w, sd.day_of_week, s) in self.x
                ]
                if week_vars:
                    self.model.Add(sum(week_vars) <= 1)

        # HC-4: Monthly cap — 1 if tagged "一个月只能1次", else 2
        for t in self._records:
            all_vars = [v for (tid, *_), v in self.x.items() if tid == t.teacher_id]
            if not all_vars:
                continue
            raw_cap = 1 if "一个月只能1次" in t.tags else 2
            # Two-system: BZ solver quota reduced by mandatory_bz_count when mandatory days exist
            if self._mandatory_bz_count > 0 and t.is_banzhuren:
                solver_cap = max(0, raw_cap - self._mandatory_bz_count)
            else:
                solver_cap = raw_cap  # original path
            self.model.Add(sum(all_vars) <= solver_cap)

        # HC-5: Weekend mutual exclusion — Friday duties + Sunday duties <= 1
        for t in self._records:
            if "一个月只能1次" in t.tags:
                continue  # already capped at 1 globally
            fri_vars = [
                self.x[(t.teacher_id, sd.week_idx, 4, s)]
                for sd in self._schedule_days if sd.day_of_week == 4
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 4, s) in self.x
            ]
            sun_vars = [
                self.x[(t.teacher_id, sd.week_idx, 6, s)]
                for sd in self._schedule_days if sd.day_of_week == 6
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 6, s) in self.x
            ]
            if fri_vars or sun_vars:
                self.model.Add(sum(fri_vars) + sum(sun_vars) <= 1)

    def _add_range_constraints(self) -> None:
        """
        HC-6: 极差约束 — for each teacher group {BZ, non-BZ}, the range
        (max − min) of NEW assignment counts must not exceed max_projected_range.
        Applies independently to total / friday / sunday new counts.
        Groups with ≤1 active teacher are skipped.
        """
        bz_active = [
            t for t in self._records
            if t.is_banzhuren and t.teacher_id in self._new_total
        ]
        non_bz_active = [
            t for t in self._records
            if not t.is_banzhuren and t.teacher_id in self._new_total
        ]

        def _add_range_for_group(
            group: list[TeacherRecord],
            new_dict: dict[int, cp_model.IntVar],
            label: str,
        ) -> None:
            if len(group) <= 1:
                return
            cap = 2
            new_vars = [new_dict[t.teacher_id] for t in group]
            max_new = self.model.NewIntVar(0, cap, f"max_new_{label}")
            min_new = self.model.NewIntVar(0, cap, f"min_new_{label}")
            self.model.AddMaxEquality(max_new, new_vars)
            self.model.AddMinEquality(min_new, new_vars)
            self.model.Add(max_new - min_new <= self._max_projected_range)

        for group_name, group in [("bz", bz_active), ("nonbz", non_bz_active)]:
            _add_range_for_group(group, self._new_total,  f"{group_name}_total")
            _add_range_for_group(group, self._new_friday, f"{group_name}_fri")
            _add_range_for_group(group, self._new_sunday, f"{group_name}_sun")

    def _set_objective(self) -> None:
        W = self._weights
        terms: list = []

        # --- Subject/day preference terms + SC-4 ---
        for (tid, w, d, slot_id), var in self.x.items():
            t = self._by_id[tid]
            if t.is_banzhuren and d in (4, 6):
                terms.append(var * W["pref_banzhuren_weekend"])
            if t.is_banzhuren and d not in (4, 6):
                # SC-4: 班主任出现在非周末（含调休周六）时惩罚，净得分为负
                terms.append(var * W["penalty_bz_non_weekend"])
            # 班主任标签优先级高于学科标签：班主任不参与学科/星期偏好加分
            if not t.is_banzhuren and t.subject == "英语" and d in (0, 3):
                terms.append(var * W["pref_english_mon_thu"])
            if not t.is_banzhuren and t.subject == "语文" and d == 1:
                terms.append(var * W["pref_chinese_tue"])
            if not t.is_banzhuren and t.subject == "政治" and d == 2:
                terms.append(var * W["pref_politics_wed"])

        # --- SC-1: pref_non_banzhuren_double ---
        for t in self._records:
            if not t.is_banzhuren and t.teacher_id in self._new_total:
                terms.append(self._new_total[t.teacher_id] * W["pref_non_banzhuren_double"])

        # --- SC-2: pref_spacing_gap ---
        gap = self._min_week_gap
        for t in self._records:
            if "一个月只能1次" in t.tags:
                continue
            tid = t.teacher_id
            for w1 in range(self._num_weeks):
                for w2 in range(w1 + gap, self._num_weeks):
                    w1_vars = [
                        self.x[(tid, w1, sd.day_of_week, s)]
                        for sd in self._schedule_days if sd.week_idx == w1
                        for s in SLOT_IDS
                        if (tid, w1, sd.day_of_week, s) in self.x
                    ]
                    w2_vars = [
                        self.x[(tid, w2, sd.day_of_week, s)]
                        for sd in self._schedule_days if sd.week_idx == w2
                        for s in SLOT_IDS
                        if (tid, w2, sd.day_of_week, s) in self.x
                    ]
                    if not w1_vars or not w2_vars:
                        continue
                    gap_ok = self.model.NewBoolVar(f"gap_{tid}_{w1}_{w2}")
                    sum_w1 = sum(w1_vars)
                    sum_w2 = sum(w2_vars)
                    self.model.Add(gap_ok <= sum_w1)
                    self.model.Add(gap_ok <= sum_w2)
                    self.model.Add(gap_ok >= sum_w1 + sum_w2 - 1)
                    terms.append(gap_ok * W["pref_spacing_gap"])

        # --- SC-3: penalty_avg_deviation ---
        penalty_weight = abs(int(W["penalty_avg_deviation"]))
        for t in self._records:
            tid = t.teacher_id
            if tid not in self._new_total:
                continue
            raw_cap = 1 if "一个月只能1次" in t.tags else 2
            # Two-system: when mandatory days exist, BZ target and cap are reduced by mandatory_bz_count
            if self._mandatory_bz_count > 0 and t.is_banzhuren:
                target = max(0, self._bz_target - self._mandatory_bz_count)
                cap = max(0, raw_cap - self._mandatory_bz_count)
            else:
                target = self._bz_target if t.is_banzhuren else self._non_bz_target
                cap = raw_cap  # original path
            max_dev = max(target, cap)
            dev_var = self.model.NewIntVar(0, max_dev, f"dev_{tid}")
            diff_lb = -target
            diff_ub = cap - target
            diff_var = self.model.NewIntVar(diff_lb, diff_ub, f"diff_{tid}")
            self.model.Add(diff_var == self._new_total[tid] - target)
            self.model.AddAbsEquality(dev_var, diff_var)
            terms.append(dev_var * (-penalty_weight))

        # --- SC-5: penalty_non_bz_weekend_double ---
        # 触发: 非班主任本月有1次周末值班(fri+sun==1) 且 总次数==2
        # wad=1 ⟺ fri+sun+total==3  →  wad<=fri+sun  AND  wad>=fri+sun+total-2
        for t in self._records:
            if t.is_banzhuren or t.teacher_id not in self._new_total:
                continue
            tid = t.teacher_id
            wad = self.model.NewBoolVar(f"wad_{tid}")
            self.model.Add(wad <= self._new_friday[tid] + self._new_sunday[tid])
            self.model.Add(
                wad >= self._new_friday[tid] + self._new_sunday[tid] + self._new_total[tid] - 2
            )
            terms.append(wad * W["penalty_non_bz_weekend_double"])

        # --- SC-6: mandatory-day BZ preference block (purely additive, original path untouched) ---
        # Activated only when mandatory_bz_count > 0.
        # Any solver-assigned BZ duty = 2nd+ overall duty (mandatory days already counted as 1st).
        # Three sub-terms per variable:
        #   · penalty_bz_second_duty       — mild penalty for any BZ solver assignment
        #   · penalty_bz_second_duty_weekend — additional larger penalty if Fri/Sun (d∈{4,6})
        #   · pref_bz_mon_thu_after_mandatory — bonus if Mon–Thu (d∈{0,1,2,3})
        # Net scores (combined with existing SC-4 / pref_banzhuren_weekend from above):
        #   Mon–Thu : SC-4(−200) + SC-6(+310) + second_duty(−30) = +80  → acceptable
        #   Fri/Sun : weekend(+100) + second_duty(−30) + second_weekend(−100) = −30 → avoided
        #   Sat(d=5): SC-4(−200) + second_duty(−30) = −230 → extremely unlikely
        if self._mandatory_bz_count > 0:
            for (tid, w, d, slot_id), var in self.x.items():
                t = self._by_id[tid]
                if not t.is_banzhuren:
                    continue
                terms.append(var * W["penalty_bz_second_duty"])
                if d in (4, 6):
                    terms.append(var * W["penalty_bz_second_duty_weekend"])
                elif d in (0, 1, 2, 3):
                    terms.append(var * W["pref_bz_mon_thu_after_mandatory"])

        if terms:
            self.model.Maximize(sum(terms))

    # -----------------------------------------------------------------------
    # Public: solve
    # -----------------------------------------------------------------------
    def solve(self) -> str:
        """
        Run CP-SAT search.
        Returns the status name string ("OPTIMAL" or "FEASIBLE").
        Raises RuntimeError if INFEASIBLE or no solution found within time limit.
        """
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = self._time_limit
        solver.parameters.num_search_workers = self._num_workers
        solver.parameters.random_seed = self._seed
        solver.parameters.log_search_progress = self._log_progress

        self.cp_solver = solver
        self.solver_status = solver.Solve(self.model)

        status_name = solver.StatusName(self.solver_status)

        if self.solver_status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            raise RuntimeError(
                f"CP-SAT returned {status_name}. "
                "Possible causes: (1) too few eligible teachers per slot, "
                "(2) monthly caps too restrictive, "
                "(3) HC-6 range constraints infeasible — history counts already differ "
                "by more than max_projected_range within a teacher group and cannot be "
                "rebalanced with the current eligible teacher pool. "
                "Check 清洗后数据 sheet and PARAMS_REGISTRY.yaml → monthly_targets."
            )
        return status_name

    # -----------------------------------------------------------------------
    # Public: post-solve verification & debug summary
    # -----------------------------------------------------------------------
    def verify_solution(self) -> list[str]:
        """
        Post-solve: verify HC-1 ~ HC-6 against solved variable values.
        Returns list of violation strings; empty = no violations.
        Must be called after a successful solve().
        """
        solver = self.cp_solver
        violations: list[str] = []

        # HC-1: Coverage
        for sd in self._schedule_days:
            # Mandatory BZ days have no solver variables — skip HC-1 check (same as model build)
            if self._mandatory_bz_count > 0 and sd.date in self._all_bz_days:
                continue
            w, d = sd.week_idx, sd.day_of_week
            for slot_id in SLOT_IDS:
                count = sum(
                    solver.Value(self.x[(tid, w, d, slot_id)])
                    for tid in self._by_id
                    if (tid, w, d, slot_id) in self.x
                )
                required = self._headcount[slot_id]
                if count != required:
                    violations.append(
                        f"HC-1 FAIL: {self._date_label(w, d)} {slot_id} "
                        f"实际{count}人，要求{required}人"
                    )

        # HC-2: No-clone
        for t in self._records:
            for sd in self._schedule_days:
                w, d = sd.week_idx, sd.day_of_week
                day_count = sum(
                    solver.Value(self.x[(t.teacher_id, w, d, s)])
                    for s in SLOT_IDS
                    if (t.teacher_id, w, d, s) in self.x
                )
                if day_count > 1:
                    violations.append(
                        f"HC-2 FAIL: {t.name} {self._date_label(w, d)} 同天被排{day_count}次"
                    )

        # HC-3: Weekly cap
        for t in self._records:
            for w in range(self._num_weeks):
                week_count = sum(
                    solver.Value(self.x[(t.teacher_id, w, sd.day_of_week, s)])
                    for sd in self._schedule_days if sd.week_idx == w
                    for s in SLOT_IDS
                    if (t.teacher_id, w, sd.day_of_week, s) in self.x
                )
                if week_count > 1:
                    violations.append(
                        f"HC-3 FAIL: {t.name} 第{w + 1}周 被排{week_count}次"
                    )

        # HC-4: Monthly cap
        for t in self._records:
            raw_cap = 1 if "一个月只能1次" in t.tags else 2
            if self._mandatory_bz_count > 0 and t.is_banzhuren:
                solver_cap = max(0, raw_cap - self._mandatory_bz_count)
            else:
                solver_cap = raw_cap  # original path
            total = sum(
                solver.Value(self.x[key])
                for key in self.x
                if key[0] == t.teacher_id
            )
            if total > solver_cap:
                violations.append(
                    f"HC-4 FAIL: {t.name} 月总次数{total}，上限{solver_cap}"
                )

        # HC-5: Weekend mutex
        for t in self._records:
            if "一个月只能1次" in t.tags:
                continue
            fri_count = sum(
                solver.Value(self.x[(t.teacher_id, sd.week_idx, 4, s)])
                for sd in self._schedule_days if sd.day_of_week == 4
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 4, s) in self.x
            )
            sun_count = sum(
                solver.Value(self.x[(t.teacher_id, sd.week_idx, 6, s)])
                for sd in self._schedule_days if sd.day_of_week == 6
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 6, s) in self.x
            )
            if fri_count + sun_count > 1:
                violations.append(
                    f"HC-5 FAIL: {t.name} 周五{fri_count}次+周日{sun_count}次 > 1"
                )

        # HC-6: Range constraint
        def _count_total(t: TeacherRecord) -> int:
            return sum(solver.Value(self.x[k]) for k in self.x if k[0] == t.teacher_id)

        def _count_fri(t: TeacherRecord) -> int:
            return sum(
                solver.Value(self.x[(t.teacher_id, sd.week_idx, 4, s)])
                for sd in self._schedule_days if sd.day_of_week == 4
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 4, s) in self.x
            )

        def _count_sun(t: TeacherRecord) -> int:
            return sum(
                solver.Value(self.x[(t.teacher_id, sd.week_idx, 6, s)])
                for sd in self._schedule_days if sd.day_of_week == 6
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 6, s) in self.x
            )

        bz_active = [
            t for t in self._records
            if t.is_banzhuren and t.teacher_id in self._new_total
        ]
        non_bz_active = [
            t for t in self._records
            if not t.is_banzhuren and t.teacher_id in self._new_total
        ]

        for group, fn, label in [
            (bz_active,     _count_total, "班主任-总次数"),
            (bz_active,     _count_fri,   "班主任-周五"),
            (bz_active,     _count_sun,   "班主任-周日"),
            (non_bz_active, _count_total, "非班主任-总次数"),
            (non_bz_active, _count_fri,   "非班主任-周五"),
            (non_bz_active, _count_sun,   "非班主任-周日"),
        ]:
            if len(group) <= 1:
                continue
            vals = [fn(t) for t in group]
            r = max(vals) - min(vals)
            if r > self._max_projected_range:
                violations.append(
                    f"HC-6 FAIL: {label} 极差={r}，超过上限{self._max_projected_range}"
                )

        return violations

    def print_solution_summary(self, output_path: str | None = None) -> None:
        """
        Print a structured debug summary: per-teacher new-assignment counts
        and the full day-by-day schedule table with real dates.
        If output_path is given, also writes to that file (UTF-8, overwrites).
        Must be called after a successful solve().
        """
        solver = self.cp_solver
        lines: list[str] = []
        add = lines.append

        add("=" * 60)
        add(f"  SOLVER DEBUG SUMMARY — {self._target_month}")
        add("=" * 60)

        def _counts(t: TeacherRecord) -> tuple[int, int, int]:
            total = sum(solver.Value(self.x[k]) for k in self.x if k[0] == t.teacher_id)
            fri = sum(
                solver.Value(self.x[(t.teacher_id, sd.week_idx, 4, s)])
                for sd in self._schedule_days if sd.day_of_week == 4
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 4, s) in self.x
            )
            sun = sum(
                solver.Value(self.x[(t.teacher_id, sd.week_idx, 6, s)])
                for sd in self._schedule_days if sd.day_of_week == 6
                for s in SLOT_IDS
                if (t.teacher_id, sd.week_idx, 6, s) in self.x
            )
            return total, fri, sun

        for group_label, group in [
            ("【班主任组】", [
                t for t in self._records
                if t.is_banzhuren and t.teacher_id in self._new_total
            ]),
            ("【非班主任组】", [
                t for t in self._records
                if not t.is_banzhuren and t.teacher_id in self._new_total
            ]),
        ]:
            add(f"\n{group_label} ({len(group)}人)")
            add(f"  {'姓名':<8} {'总次数':>4} {'周五':>4} {'周日':>4}  备注")
            add(f"  {'-' * 8} {'-' * 4} {'-' * 4} {'-' * 4}")
            totals, fris, suns = [], [], []
            for t in sorted(group, key=lambda x: x.name):
                total, fri, sun = _counts(t)
                totals.append(total)
                fris.append(fri)
                suns.append(sun)
                note = "*** 被排2次" if (total == 2 and t.is_banzhuren) else ""
                add(f"  {t.name:<8} {total:>4} {fri:>4} {sun:>4}  {note}")
            if totals:
                add(
                    f"  {'[极差]':<8} {max(totals) - min(totals):>4} "
                    f"{max(fris) - min(fris):>4} {max(suns) - min(suns):>4}"
                )

        add(f"\n{'=' * 60}")
        add(f"  排班明细（{self._target_month}）")
        add("=" * 60)
        unassigned = self._config["output"]["unassigned_placeholder"]
        for w in range(self._num_weeks):
            add(f"\n  第{w + 1}周:")
            week_days = [sd for sd in self._schedule_days if sd.week_idx == w]
            for sd in week_days:
                label = self._date_label(sd.week_idx, sd.day_of_week)
                parts = [f"    {label}"]
                for slot_id in SLOT_IDS:
                    zone = SLOT_ZONE[slot_id]
                    name = unassigned
                    for tid, t in self._by_id.items():
                        key = (tid, sd.week_idx, sd.day_of_week, slot_id)
                        if key in self.x and solver.Value(self.x[key]) == 1:
                            name = t.name
                            break
                    parts.append(f"{zone}:{name}")
                add("  ".join(parts))

        add("\n" + "=" * 60)

        output = "\n".join(lines)
        print(output)

        if output_path:
            import pathlib as _pl
            _pl.Path(output_path).write_text(output, encoding="utf-8")
            print(f"\n[debug report → {output_path}]")

    # -----------------------------------------------------------------------
    # Public: export
    # -----------------------------------------------------------------------
    def export_to_excel(self, output_path: str) -> None:
        """
        Write the solved schedule as a new sheet into *output_path*.
        Sheet name: e.g. "5月排班（暂定）"
        Date labels: e.g. "5月6日(周二)"
        Replaces the sheet if it already exists.
        Must be called after solve().
        """
        if self.cp_solver is None or self.solver_status not in (
            cp_model.OPTIMAL, cp_model.FEASIBLE
        ):
            raise RuntimeError(
                "export_to_excel() called before a successful solve(). "
                "Call solve() first."
            )

        rows: list[dict] = []
        unassigned_placeholder: str = self._config["output"]["unassigned_placeholder"]
        # Track Excel row numbers (1-based; row 1 = header, data starts at row 2)
        # for mandatory BZ days so we can merge B-D after writing.
        mandatory_row_indices: list[int] = []

        for excel_row, sd in enumerate(self._schedule_days, start=2):
            w, d = sd.week_idx, sd.day_of_week
            label = self._date_label(w, d)
            # Two-system: mandatory days get a placeholder row; normal days get individual assignments
            if self._mandatory_bz_count > 0 and sd.date in self._all_bz_days:
                row: dict[str, str] = {"日期": label, "1楼": "全体班主任", "2-3楼": "", "4-5楼": ""}
                mandatory_row_indices.append(excel_row)
            else:
                row = {"日期": label}
                for slot_id in SLOT_IDS:
                    zone = SLOT_ZONE[slot_id]
                    assigned_name = unassigned_placeholder
                    for tid, t in self._by_id.items():
                        if (tid, w, d, slot_id) in self.x:
                            if self.cp_solver.Value(self.x[(tid, w, d, slot_id)]) == 1:
                                assigned_name = t.name
                                break
                    row[zone] = assigned_name
            rows.append(row)

        df = pd.DataFrame(rows, columns=["日期", "1楼", "2-3楼", "4-5楼"])

        banzhuren_names: set[str] = {t.name for t in self._records if t.is_banzhuren}
        orange_yellow = PatternFill(fill_type="solid", fgColor="FFB300")

        sheet_name = self._output_sheet_name
        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            # Row 1 is the header; data starts at row 2
            # Columns: A=日期, B=1楼, C=2-3楼, D=4-5楼
            for row_idx in range(2, len(rows) + 2):
                for col_idx in range(2, 5):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value in banzhuren_names:
                        cell.fill = orange_yellow

            # Merge B-D for mandatory BZ days and apply "全体班主任" with orange-yellow fill.
            # Must run after the BZ individual-cell fill loop (mandatory rows are not in
            # banzhuren_names so the loop above is a no-op for them; this pass overrides correctly).
            for r in mandatory_row_indices:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
                merged_cell = ws.cell(row=r, column=2)
                merged_cell.value = "全体班主任"
                merged_cell.fill = orange_yellow
                merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        del df


# ---------------------------------------------------------------------------
# Verification entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    excel_path = sys.argv[1] if len(sys.argv) > 1 else _DEFAULT_EXCEL_PATH

    # Determine target month: command-line arg, state file, or today
    if len(sys.argv) > 2:
        target_month = sys.argv[2]
    elif _STATE_PATH.exists():
        _state = json.loads(_STATE_PATH.read_text(encoding="utf-8"))
        target_month = _next_month(_state["last_generated"])
    else:
        _today = datetime.date.today()
        target_month = f"{_today.year:04d}-{_today.month:02d}"

    # Load special dates for this month
    special_dates_for_month: dict = {}
    if _SPECIAL_DATES_PATH.exists():
        _all_special = json.loads(_SPECIAL_DATES_PATH.read_text(encoding="utf-8"))
        special_dates_for_month = {
            k: v for k, v in _all_special.items() if not k.startswith("_")
        }.get(target_month, {})

    print(f"Excel: {excel_path}")
    print(f"Target month: {target_month}")
    if special_dates_for_month:
        print(f"Special dates: {special_dates_for_month.get('notes', '')}")

    print("Loading teacher records...")
    records = load_teacher_records(excel_path)
    print(f"  {len(records)} records loaded.")

    print("Building CP-SAT model...")
    ds = DutySolver(records, _CFG, target_month, special_dates_for_month)
    print(
        f"  {len(ds._schedule_days)} duty days  |  "
        f"{ds._num_weeks} weeks  |  "
        f"{len(ds.x)} decision variables."
    )

    print("Solving...")
    status = ds.solve()
    print(f"  Solver status: {status}")

    debug_path = str(pathlib.Path(__file__).parent / "debug_solver_run.txt")
    violations = ds.verify_solution()
    if violations:
        print(f"\n[!] {len(violations)} constraint violation(s) found:")
        for v in violations:
            print(f"    {v}")
    else:
        print("\n[OK] All hard constraints verified — no violations.")

    ds.print_solution_summary(output_path=debug_path)

    print("Exporting to Excel...")
    ds.export_to_excel(excel_path)
    print(f"  结果已写入 '{ds._output_sheet_name}' sheet。")
